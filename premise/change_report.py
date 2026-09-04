"""Structured, exact source-to-scenario change reports.

The Excel artifact in this module is intentionally a compact review surface.
Every reportable row is written to the companion Parquet file while summaries
are accumulated in the same ordered pass.
"""

from __future__ import annotations

import hashlib
import heapq
import json
import math
import os
import re
import threading
import uuid
from collections import Counter, defaultdict, deque
from dataclasses import dataclass, field
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Any, Iterable, Iterator, Literal, Mapping, Sequence

import openpyxl
import pyarrow as pa
import pyarrow.parquet as pq
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from .inventory_store import InventoryStore
from .provenance import PROVENANCE_SCHEMA_VERSION, ProvenanceEvent
from .validation_framework import ValidationReport

REPORT_SCHEMA_VERSION = 2
DETAIL_BATCH_SIZE = 4096

IGNORED_FIELDS = frozenset(
    {
        "database",
        "output",
        "id",
        "activity id",
        "activity_id",
        "exchange id",
        "exchange_id",
        "log parameters",
        "applied functions",
    }
)
UNCERTAINTY_FIELDS = frozenset(
    {
        "uncertainty type",
        "loc",
        "scale",
        "shape",
        "minimum",
        "maximum",
        "negative",
    }
)
EXCHANGE_IDENTITY_FIELDS = frozenset(
    {"type", "name", "product", "reference product", "unit"}
)
EXCHANGE_PROVIDER_FIELDS = frozenset({"input", "location", "code", "database"})


@dataclass(frozen=True, slots=True)
class ChangeReportArtifacts:
    """Paths and immutable identities produced by a V2 report session."""

    report_id: str
    status: Literal["passed", "failed"]
    workbook_path: Path
    details_path: Path | None
    scenario_identities: tuple[tuple[Any, ...], ...]
    source_fingerprint: str
    validation_certificate_keys: tuple[str, ...]

    def __post_init__(self) -> None:
        object.__setattr__(self, "workbook_path", Path(self.workbook_path))
        if self.details_path is not None:
            object.__setattr__(self, "details_path", Path(self.details_path))
        object.__setattr__(
            self,
            "scenario_identities",
            tuple(tuple(identity) for identity in self.scenario_identities),
        )
        object.__setattr__(
            self, "validation_certificate_keys", tuple(self.validation_certificate_keys)
        )


@dataclass(frozen=True, slots=True)
class ReportScenario:
    identity: tuple[Any, ...]
    store: InventoryStore
    validation_report: ValidationReport | None = None
    provenance_payload: Mapping[str, Any] | None = None
    definition: Mapping[str, Any] | None = None

    def __post_init__(self) -> None:
        object.__setattr__(self, "identity", tuple(self.identity))


@dataclass(frozen=True, slots=True)
class Attribution:
    sector: str
    transformation: str
    reason_code: str
    explanation: str
    iam_variable: str | None = None
    algorithm: str | None = None
    configuration_reference: str | None = None
    proxy: str | None = None
    fallback_rank: int | None = None


@dataclass(slots=True)
class _ActivityLocator:
    activity_id: int
    code: str | None
    semantic_key: tuple[str, str, str, str]
    visible_hash: str
    occurrence: int = 0


@dataclass(slots=True)
class _Summary:
    source_activity_count: int = 0
    source_exchange_count: int = 0
    final_counts: dict[str, tuple[int, int]] = field(default_factory=dict)
    scenario_counts: dict[str, Counter] = field(
        default_factory=lambda: defaultdict(Counter)
    )
    sector_counts: dict[tuple[str, str, str], Counter] = field(
        default_factory=lambda: defaultdict(Counter)
    )
    key_heaps: dict[tuple[str, str, str, str], list] = field(
        default_factory=lambda: defaultdict(list)
    )
    market_rows: list[dict[str, Any]] = field(default_factory=list)
    fallback_counts: Counter = field(default_factory=Counter)
    fallback_payloads: dict[tuple[Any, ...], dict[str, Any]] = field(
        default_factory=dict
    )
    methodology_rows: list[dict[str, Any]] = field(default_factory=list)
    final_fingerprints: dict[str, str] = field(default_factory=dict)
    _heap_counter: int = 0

    def add_fallback(self, row: Mapping[str, Any]) -> None:
        key = (
            row.get("scenario"),
            row.get("requested geography"),
            row.get("requested technology"),
            row.get("selected proxy"),
            row.get("fallback rank"),
            row.get("explanation"),
        )
        self.fallback_counts[key] += int(row.get("affected activity count", 1))
        self.fallback_payloads[key] = dict(row)

    def fallback_rows(self) -> list[dict[str, Any]]:
        rows = []
        for key in sorted(self.fallback_counts, key=str):
            row = dict(self.fallback_payloads[key])
            row["affected activity count"] = self.fallback_counts[key]
            rows.append(row)
        return rows

    def consume_activity(self, records: list[dict[str, Any]]) -> None:
        if not records:
            return
        scenario = records[0]["scenario_identity"]
        seen_scenario: set[tuple[Any, ...]] = set()
        seen_sector: set[tuple[Any, ...]] = set()
        for record in records:
            category = _summary_category(record)
            object_token = _object_token(record, category)
            scenario_token = (object_token, category)
            if scenario_token not in seen_scenario:
                self.scenario_counts[scenario][category] += 1
                seen_scenario.add(scenario_token)
            for sector, transformation in record.pop(
                "_sector_transformations", (("unattributed", "unattributed"),)
            ):
                sector_token = (sector, transformation, object_token, category)
                if sector_token not in seen_sector:
                    self.sector_counts[(scenario, sector, transformation)][
                        category
                    ] += 1
                    seen_sector.add(sector_token)
                self._consider_key_change(record, sector, transformation)

    def _consider_key_change(
        self, record: Mapping[str, Any], sector: str, transformation: str
    ) -> None:
        delta = record.get("absolute_delta")
        if delta is None or not math.isfinite(delta):
            return
        old = record.get("old_numeric")
        relative = record.get("relative_delta")
        score = abs(delta) if old in (None, 0) else abs(relative or 0.0)
        key = (
            record["scenario_identity"],
            sector,
            record["object_type"],
            record["change_type"],
        )
        self._heap_counter += 1
        row = {
            "scenario": record["scenario_identity"],
            "sector": sector,
            "transformation": transformation,
            "object type": record["object_type"],
            "change type": record["change_type"],
            "activity": record.get("activity_name"),
            "product": record.get("activity_product"),
            "location": record.get("activity_location"),
            "exchange": record.get("exchange_name"),
            "field": record.get("changed_field"),
            "old value": record.get("old_numeric"),
            "new value": record.get("new_numeric"),
            "absolute delta": delta,
            "relative delta": relative,
            "unit": record.get("unit"),
            "reason": record.get("explanation"),
        }
        item = (score, self._heap_counter, row)
        heap = self.key_heaps[key]
        if len(heap) < 20:
            heapq.heappush(heap, item)
        elif item[:2] > heap[0][:2]:
            heapq.heapreplace(heap, item)

    def key_change_rows(self) -> list[dict[str, Any]]:
        rows = []
        for key in sorted(self.key_heaps):
            rows.extend(
                item[2]
                for item in sorted(
                    self.key_heaps[key], key=lambda entry: (-entry[0], entry[1])
                )
            )
        return rows


@dataclass(frozen=True, slots=True)
class ChangeReportCacheEntry:
    cache_key: tuple[Any, ...]
    report_id: str
    details_path: Path
    summary: _Summary


@dataclass(frozen=True, slots=True)
class GeneratedChangeReport:
    artifacts: ChangeReportArtifacts
    cache_entry: ChangeReportCacheEntry


DETAIL_SCHEMA = pa.schema(
    [
        pa.field("report_schema_version", pa.int16(), nullable=False),
        pa.field("report_id", pa.string(), nullable=False),
        pa.field("build_id", pa.string(), nullable=False),
        pa.field("scenario_order", pa.int32(), nullable=False),
        pa.field("scenario_identity", pa.string(), nullable=False),
        pa.field("source_fingerprint", pa.string(), nullable=False),
        pa.field("final_fingerprint", pa.string(), nullable=False),
        pa.field("validation_certificate_key", pa.string()),
        pa.field("transformations", pa.list_(pa.string()), nullable=False),
        pa.field("sector", pa.string(), nullable=False),
        pa.field("object_type", pa.string(), nullable=False),
        pa.field("change_type", pa.string(), nullable=False),
        pa.field("activity_code", pa.string()),
        pa.field("activity_name", pa.string()),
        pa.field("activity_product", pa.string()),
        pa.field("activity_location", pa.string()),
        pa.field("activity_unit", pa.string()),
        pa.field("activity_occurrence", pa.int32(), nullable=False),
        pa.field("exchange_type", pa.string()),
        pa.field("exchange_name", pa.string()),
        pa.field("exchange_product", pa.string()),
        pa.field("exchange_location", pa.string()),
        pa.field("exchange_occurrence", pa.int32()),
        pa.field("old_provider_identity", pa.string()),
        pa.field("new_provider_identity", pa.string()),
        pa.field("changed_field", pa.string()),
        pa.field("old_value_json", pa.string()),
        pa.field("new_value_json", pa.string()),
        pa.field("old_numeric", pa.float64()),
        pa.field("new_numeric", pa.float64()),
        pa.field("absolute_delta", pa.float64()),
        pa.field("relative_delta", pa.float64()),
        pa.field("unit", pa.string()),
        pa.field("reason_code", pa.string(), nullable=False),
        pa.field("explanation", pa.string(), nullable=False),
        pa.field("iam_variable", pa.string()),
        pa.field("algorithm", pa.string()),
        pa.field("configuration_reference", pa.string()),
        pa.field("proxy", pa.string()),
        pa.field("fallback_rank", pa.int32()),
    ],
    metadata={
        b"premise_report_schema_version": str(REPORT_SCHEMA_VERSION).encode("ascii"),
        b"format": b"premise-structured-change-audit",
    },
)


def _plain(value: Any) -> Any:
    if isinstance(value, Mapping):
        return {
            str(key): _plain(item)
            for key, item in sorted(value.items(), key=lambda item: str(item[0]))
            if not str(key).startswith("_") and key not in IGNORED_FIELDS
        }
    if isinstance(value, (list, tuple)):
        return [_plain(item) for item in value]
    if isinstance(value, (set, frozenset)):
        return sorted((_plain(item) for item in value), key=_canonical_json)
    if isinstance(value, bytes):
        return {"__bytes__": value.hex()}
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, Path):
        return str(value)
    item = getattr(value, "item", None)
    if callable(item):
        try:
            return _plain(item())
        except (TypeError, ValueError):
            pass
    if isinstance(value, float):
        if math.isnan(value):
            return {"__float__": "NaN"}
        if math.isinf(value):
            return {"__float__": "Infinity" if value > 0 else "-Infinity"}
    if isinstance(value, (str, int, float, bool)) or value is None:
        return value
    return {"__type__": type(value).__name__, "value": str(value)}


def _canonical_json(value: Any) -> str:
    return json.dumps(
        _plain(value), sort_keys=True, separators=(",", ":"), ensure_ascii=False
    )


def _stable_hash(value: Any) -> str:
    return hashlib.sha256(_canonical_json(value).encode("utf-8")).hexdigest()


def _visible_mapping(
    payload: Mapping[str, Any], *, ignored: Iterable[str] = ()
) -> dict[str, Any]:
    ignored_fields = set(ignored) | set(IGNORED_FIELDS)
    return {
        str(key): _plain(value)
        for key, value in payload.items()
        if not str(key).startswith("_") and key not in ignored_fields
    }


def _visible_activity(payload: Mapping[str, Any]) -> dict[str, Any]:
    return _visible_mapping(payload, ignored={"exchanges", "code"})


def _visible_exchange(payload: Mapping[str, Any]) -> dict[str, Any]:
    return _visible_mapping(payload)


def _semantic_key(payload: Mapping[str, Any]) -> tuple[str, str, str, str]:
    return (
        str(payload.get("name") or ""),
        str(payload.get("reference product", payload.get("product")) or ""),
        str(payload.get("location") or ""),
        str(payload.get("unit") or ""),
    )


def _scenario_label(identity: Sequence[Any]) -> str:
    return _canonical_json(list(identity))


def _number(value: Any) -> float | None:
    if isinstance(value, bool):
        return None
    item = getattr(value, "item", None)
    if callable(item):
        try:
            value = item()
        except (TypeError, ValueError):
            return None
    if isinstance(value, (int, float)):
        return float(value)
    return None


def _numeric_change(old: Any, new: Any) -> tuple[float | None, ...]:
    old_number = _number(old)
    new_number = _number(new)
    if old_number is None or new_number is None:
        return old_number, new_number, None, None
    delta = new_number - old_number
    if not math.isfinite(delta):
        return old_number, new_number, delta, None
    relative = None
    if old_number != 0 and math.isfinite(old_number):
        relative = delta / abs(old_number)
    return old_number, new_number, delta, relative


def _activity_index(
    store: InventoryStore,
) -> tuple[list[_ActivityLocator], int, str]:
    locators = []
    exchange_count = 0
    digest = hashlib.sha256()
    for activity_id in store.iter_activity_ids():
        payload = store.activity(activity_id).to_dict()
        exchanges = payload.get("exchanges", ())
        exchange_count += len(exchanges)
        visible = {
            "activity": _visible_activity(payload),
            "exchanges": [_visible_exchange(exchange) for exchange in exchanges],
        }
        visible_hash = _stable_hash(visible)
        digest.update(visible_hash.encode("ascii"))
        code = payload.get("code")
        locators.append(
            _ActivityLocator(
                activity_id=int(activity_id),
                code=None if code in (None, "") else str(code),
                semantic_key=_semantic_key(payload),
                visible_hash=visible_hash,
            )
        )
    by_semantic: dict[tuple[str, str, str, str], list[_ActivityLocator]] = defaultdict(
        list
    )
    for locator in locators:
        by_semantic[locator.semantic_key].append(locator)
    for group in by_semantic.values():
        for occurrence, locator in enumerate(group):
            locator.occurrence = occurrence
    return locators, exchange_count, digest.hexdigest()


def _pair_activities(
    source: list[_ActivityLocator], final: list[_ActivityLocator]
) -> list[tuple[_ActivityLocator | None, _ActivityLocator | None]]:
    pairs: list[tuple[_ActivityLocator | None, _ActivityLocator | None]] = []
    source_remaining = {locator.activity_id: locator for locator in source}
    final_remaining = {locator.activity_id: locator for locator in final}
    source_codes: dict[str, list[_ActivityLocator]] = defaultdict(list)
    final_codes: dict[str, list[_ActivityLocator]] = defaultdict(list)
    for locator in source:
        if locator.code is not None:
            source_codes[locator.code].append(locator)
    for locator in final:
        if locator.code is not None:
            final_codes[locator.code].append(locator)
    for code in sorted(source_codes.keys() & final_codes.keys()):
        old_group = sorted(
            source_codes[code], key=lambda item: (item.occurrence, item.activity_id)
        )
        new_group = sorted(
            final_codes[code], key=lambda item: (item.occurrence, item.activity_id)
        )
        for old, new in zip(old_group, new_group):
            pairs.append((old, new))
            source_remaining.pop(old.activity_id, None)
            final_remaining.pop(new.activity_id, None)

    old_semantics: dict[tuple[str, str, str, str], list[_ActivityLocator]] = (
        defaultdict(list)
    )
    new_semantics: dict[tuple[str, str, str, str], list[_ActivityLocator]] = (
        defaultdict(list)
    )
    for locator in source_remaining.values():
        old_semantics[locator.semantic_key].append(locator)
    for locator in final_remaining.values():
        new_semantics[locator.semantic_key].append(locator)
    for semantic in sorted(old_semantics.keys() & new_semantics.keys()):
        old_group = sorted(
            old_semantics[semantic],
            key=lambda item: (item.occurrence, item.activity_id),
        )
        new_group = sorted(
            new_semantics[semantic],
            key=lambda item: (item.occurrence, item.activity_id),
        )
        old_eligible = [item for item in old_group if item.code is None]
        new_eligible = [item for item in new_group if item.code is None]
        if old_eligible:
            new_eligible = new_group
        elif new_eligible:
            old_eligible = old_group
        else:
            continue
        for old, new in zip(old_eligible, new_eligible):
            if (
                old.activity_id not in source_remaining
                or new.activity_id not in final_remaining
            ):
                continue
            pairs.append((old, new))
            source_remaining.pop(old.activity_id, None)
            final_remaining.pop(new.activity_id, None)

    pairs.extend((locator, None) for locator in source_remaining.values())
    pairs.extend((None, locator) for locator in final_remaining.values())

    def sort_key(pair):
        locator = pair[1] or pair[0]
        return (
            locator.semantic_key,
            locator.code or "",
            locator.occurrence,
            pair[0] is None,
            pair[1] is None,
        )

    return sorted(pairs, key=sort_key)


def _provider_identity(exchange: Mapping[str, Any]) -> dict[str, Any] | None:
    if exchange.get("type") != "technosphere":
        return None
    exchange_input = exchange.get("input")
    input_code = None
    if isinstance(exchange_input, (tuple, list)) and len(exchange_input) == 2:
        input_code = exchange_input[1]
    return {
        "code": input_code or exchange.get("code"),
        "name": exchange.get("name"),
        "product": exchange.get("product", exchange.get("reference product")),
        "location": exchange.get("location"),
    }


def _exchange_group_key(exchange: Mapping[str, Any]) -> tuple[str, str, str, str]:
    return (
        str(exchange.get("type") or ""),
        str(exchange.get("name") or ""),
        str(exchange.get("product", exchange.get("reference product")) or ""),
        str(exchange.get("unit") or ""),
    )


def _exchange_metadata(exchange: Mapping[str, Any]) -> dict[str, Any]:
    return {
        key: value
        for key, value in _visible_exchange(exchange).items()
        if key
        not in EXCHANGE_IDENTITY_FIELDS
        | EXCHANGE_PROVIDER_FIELDS
        | UNCERTAINTY_FIELDS
        | {"amount"}
    }


def _cancel_exact(
    old: list[Mapping[str, Any]], new: list[Mapping[str, Any]]
) -> tuple[list[Mapping[str, Any]], list[Mapping[str, Any]]]:
    new_by_signature: dict[str, deque[int]] = defaultdict(deque)
    for index, exchange in enumerate(new):
        new_by_signature[_canonical_json(_visible_exchange(exchange))].append(index)
    used_new = set()
    old_remaining = []
    for exchange in old:
        signature = _canonical_json(_visible_exchange(exchange))
        candidates = new_by_signature.get(signature)
        if candidates:
            used_new.add(candidates.popleft())
        else:
            old_remaining.append(exchange)
    return old_remaining, [
        item for index, item in enumerate(new) if index not in used_new
    ]


def _pair_exchange_group(
    old: list[Mapping[str, Any]], new: list[Mapping[str, Any]]
) -> tuple[
    list[tuple[Mapping[str, Any], Mapping[str, Any], int]],
    list[tuple[Mapping[str, Any], int]],
    list[tuple[Mapping[str, Any], int]],
]:
    old_sorted = sorted(old, key=lambda item: _canonical_json(_visible_exchange(item)))
    new_sorted = sorted(new, key=lambda item: _canonical_json(_visible_exchange(item)))
    edges = []
    for old_index, old_exchange in enumerate(old_sorted):
        old_provider = _canonical_json(_provider_identity(old_exchange))
        old_amount = _number(old_exchange.get("amount"))
        old_metadata = _stable_hash(_exchange_metadata(old_exchange))
        for new_index, new_exchange in enumerate(new_sorted):
            new_provider = _canonical_json(_provider_identity(new_exchange))
            new_amount = _number(new_exchange.get("amount"))
            if (
                old_amount is None
                or new_amount is None
                or not math.isfinite(old_amount)
                or not math.isfinite(new_amount)
            ):
                distance = math.inf
            else:
                distance = abs(new_amount - old_amount)
            edges.append(
                (
                    old_provider != new_provider,
                    old_exchange.get("location") != new_exchange.get("location"),
                    distance,
                    old_metadata != _stable_hash(_exchange_metadata(new_exchange)),
                    old_metadata,
                    _stable_hash(_exchange_metadata(new_exchange)),
                    old_index,
                    new_index,
                )
            )
    paired_old = set()
    paired_new = set()
    pairs = []
    for *_, old_index, new_index in sorted(edges):
        if old_index in paired_old or new_index in paired_new:
            continue
        paired_old.add(old_index)
        paired_new.add(new_index)
        pairs.append((old_sorted[old_index], new_sorted[new_index], new_index))
        if len(pairs) == min(len(old_sorted), len(new_sorted)):
            break
    removals = [
        (exchange, index)
        for index, exchange in enumerate(old_sorted)
        if index not in paired_old
    ]
    additions = [
        (exchange, index)
        for index, exchange in enumerate(new_sorted)
        if index not in paired_new
    ]
    return sorted(pairs, key=lambda item: item[2]), removals, additions


class _AttributionIndex:
    def __init__(self, scenario: ReportScenario) -> None:
        self.by_code: dict[str, list[Attribution]] = defaultdict(list)
        self.by_semantic: dict[tuple[str, str, str], list[Attribution]] = defaultdict(
            list
        )
        self.events: list[ProvenanceEvent] = []
        payload = scenario.provenance_payload or {}
        for item in payload.get("events", ()) if isinstance(payload, Mapping) else ():
            try:
                event = ProvenanceEvent.from_dict(item)
            except (KeyError, TypeError, ValueError):
                continue
            self.events.append(event)
            attribution = Attribution(
                sector=event.sector,
                transformation=event.transformation,
                reason_code=event.reason_code,
                explanation=event.explanation,
                iam_variable=event.iam_variable,
                algorithm=event.algorithm,
                configuration_reference=event.configuration_reference,
                proxy=event.proxy,
                fallback_rank=event.fallback_rank,
            )
            self._add(event.activity, attribution)

        definition = scenario.definition or {}
        intents = definition.get("_validation_intents", {})
        for intent in intents.values() if isinstance(intents, Mapping) else ():
            if not isinstance(intent, Mapping):
                continue
            transformation = str(intent.get("transformation") or "unattributed")
            attribution = Attribution(
                sector=transformation,
                transformation=transformation,
                reason_code=f"{transformation}.validation_intent",
                explanation=(
                    f"Change falls within the certified {transformation} "
                    "transformation scope."
                ),
                algorithm=intent.get("algorithm"),
            )
            for key in intent.get("affected_activity_keys", ()):
                if isinstance(key, (tuple, list)) and len(key) >= 3:
                    self.by_semantic[
                        tuple(str(value or "") for value in key[:3])
                    ].append(attribution)

    def _add(self, activity: Mapping[str, Any], attribution: Attribution) -> None:
        code = activity.get("code")
        if code not in (None, ""):
            self.by_code[str(code)].append(attribution)
        key = (
            str(activity.get("name") or ""),
            str(activity.get("product", activity.get("reference product")) or ""),
            str(activity.get("location") or ""),
        )
        self.by_semantic[key].append(attribution)

    def lookup(self, activity: Mapping[str, Any]) -> tuple[Attribution, ...]:
        candidates = []
        code = activity.get("code")
        if code not in (None, ""):
            candidates.extend(self.by_code.get(str(code), ()))
        key = (
            str(activity.get("name") or ""),
            str(activity.get("reference product", activity.get("product")) or ""),
            str(activity.get("location") or ""),
        )
        candidates.extend(self.by_semantic.get(key, ()))
        unique = {}
        for item in candidates:
            signature = (
                item.sector,
                item.transformation,
                item.reason_code,
                item.explanation,
            )
            unique[signature] = item
        if unique:
            return tuple(unique[key] for key in sorted(unique))
        return (
            Attribution(
                sector="unattributed",
                transformation="unattributed",
                reason_code="unattributed",
                explanation=(
                    "No transformation provenance or certified scope declaration "
                    "matched this change."
                ),
            ),
        )


def _record_base(
    *,
    report_id: str,
    build_id: str,
    scenario_order: int,
    scenario_label: str,
    source_fingerprint: str,
    final_fingerprint: str,
    certificate_key: str | None,
    activity: Mapping[str, Any],
    occurrence: int,
    attributions: Sequence[Attribution],
) -> dict[str, Any]:
    transformations = sorted({item.transformation for item in attributions})
    sectors = sorted({item.sector for item in attributions})

    def joined(attribute: str) -> str | None:
        values = sorted(
            {
                str(value)
                for value in (getattr(item, attribute) for item in attributions)
                if value not in (None, "")
            }
        )
        return " | ".join(values) if values else None

    return {
        "report_schema_version": REPORT_SCHEMA_VERSION,
        "report_id": report_id,
        "build_id": build_id,
        "scenario_order": scenario_order,
        "scenario_identity": scenario_label,
        "source_fingerprint": source_fingerprint,
        "final_fingerprint": final_fingerprint,
        "validation_certificate_key": certificate_key,
        "transformations": transformations,
        "sector": ", ".join(sectors),
        "activity_code": (
            None if activity.get("code") in (None, "") else str(activity.get("code"))
        ),
        "activity_name": activity.get("name"),
        "activity_product": activity.get("reference product", activity.get("product")),
        "activity_location": activity.get("location"),
        "activity_unit": activity.get("unit"),
        "activity_occurrence": occurrence,
        "exchange_type": None,
        "exchange_name": None,
        "exchange_product": None,
        "exchange_location": None,
        "exchange_occurrence": None,
        "old_provider_identity": None,
        "new_provider_identity": None,
        "changed_field": None,
        "old_value_json": None,
        "new_value_json": None,
        "old_numeric": None,
        "new_numeric": None,
        "absolute_delta": None,
        "relative_delta": None,
        "unit": activity.get("unit"),
        "reason_code": joined("reason_code") or "unattributed",
        "explanation": " | ".join(
            dict.fromkeys(item.explanation for item in attributions)
        ),
        "iam_variable": joined("iam_variable"),
        "algorithm": joined("algorithm"),
        "configuration_reference": joined("configuration_reference"),
        "proxy": joined("proxy"),
        "fallback_rank": next(
            (
                item.fallback_rank
                for item in attributions
                if item.fallback_rank is not None
            ),
            None,
        ),
        "_sector_transformations": tuple(
            sorted({(item.sector, item.transformation) for item in attributions})
        ),
    }


def _value_record(
    base: Mapping[str, Any],
    *,
    object_type: str,
    change_type: str,
    changed_field: str | None,
    old: Any,
    new: Any,
    unit: str | None = None,
) -> dict[str, Any]:
    record = dict(base)
    old_numeric, new_numeric, delta, relative = _numeric_change(old, new)
    record.update(
        {
            "object_type": object_type,
            "change_type": change_type,
            "changed_field": changed_field,
            "old_value_json": None if old is None else _canonical_json(old),
            "new_value_json": None if new is None else _canonical_json(new),
            "old_numeric": old_numeric,
            "new_numeric": new_numeric,
            "absolute_delta": delta,
            "relative_delta": relative,
            "unit": unit or base.get("unit"),
        }
    )
    return record


def _exchange_base(
    base: Mapping[str, Any], exchange: Mapping[str, Any], occurrence: int
) -> dict[str, Any]:
    record = dict(base)
    record.update(
        {
            "exchange_type": exchange.get("type"),
            "exchange_name": exchange.get("name"),
            "exchange_product": exchange.get(
                "product", exchange.get("reference product")
            ),
            "exchange_location": exchange.get("location"),
            "exchange_occurrence": occurrence,
            "unit": exchange.get("unit"),
        }
    )
    return record


def _activity_records(
    source_store: InventoryStore,
    final_store: InventoryStore,
    pair: tuple[_ActivityLocator | None, _ActivityLocator | None],
    *,
    report_id: str,
    build_id: str,
    scenario_order: int,
    scenario_label: str,
    source_fingerprint: str,
    final_fingerprint: str,
    certificate_key: str | None,
    attribution_index: _AttributionIndex,
    summary: _Summary,
) -> list[dict[str, Any]]:
    old_locator, new_locator = pair
    old_activity = (
        source_store.activity(old_locator.activity_id).to_dict()
        if old_locator is not None
        else None
    )
    new_activity = (
        final_store.activity(new_locator.activity_id).to_dict()
        if new_locator is not None
        else None
    )
    identity_activity = new_activity or old_activity
    occurrence = (new_locator or old_locator).occurrence
    attributions = attribution_index.lookup(identity_activity)
    base = _record_base(
        report_id=report_id,
        build_id=build_id,
        scenario_order=scenario_order,
        scenario_label=scenario_label,
        source_fingerprint=source_fingerprint,
        final_fingerprint=final_fingerprint,
        certificate_key=certificate_key,
        activity=identity_activity,
        occurrence=occurrence,
        attributions=attributions,
    )
    records: list[dict[str, Any]] = []
    if old_activity is None:
        records.append(
            _value_record(
                base,
                object_type="activity",
                change_type="addition",
                changed_field=None,
                old=None,
                new=_visible_activity(new_activity),
            )
        )
        for exchange_occurrence, exchange in enumerate(
            sorted(
                new_activity.get("exchanges", ()),
                key=lambda item: (
                    _exchange_group_key(item),
                    _canonical_json(_visible_exchange(item)),
                ),
            )
        ):
            exchange_record = _exchange_base(base, exchange, exchange_occurrence)
            records.append(
                _value_record(
                    exchange_record,
                    object_type="exchange",
                    change_type="addition",
                    changed_field=None,
                    old=None,
                    new=_visible_exchange(exchange),
                    unit=exchange.get("unit"),
                )
            )
        return records
    if new_activity is None:
        records.append(
            _value_record(
                base,
                object_type="activity",
                change_type="removal",
                changed_field=None,
                old=_visible_activity(old_activity),
                new=None,
            )
        )
        for exchange_occurrence, exchange in enumerate(
            sorted(
                old_activity.get("exchanges", ()),
                key=lambda item: (
                    _exchange_group_key(item),
                    _canonical_json(_visible_exchange(item)),
                ),
            )
        ):
            exchange_record = _exchange_base(base, exchange, exchange_occurrence)
            records.append(
                _value_record(
                    exchange_record,
                    object_type="exchange",
                    change_type="removal",
                    changed_field=None,
                    old=_visible_exchange(exchange),
                    new=None,
                    unit=exchange.get("unit"),
                )
            )
        return records

    old_visible = _visible_activity(old_activity)
    new_visible = _visible_activity(new_activity)
    for field_name in sorted(old_visible.keys() | new_visible.keys()):
        old_value = old_visible.get(field_name)
        new_value = new_visible.get(field_name)
        if _canonical_json(old_value) == _canonical_json(new_value):
            continue
        records.append(
            _value_record(
                base,
                object_type="activity",
                change_type="field modification",
                changed_field=field_name,
                old=old_value,
                new=new_value,
            )
        )

    old_exchanges, new_exchanges = _cancel_exact(
        list(old_activity.get("exchanges", ())),
        list(new_activity.get("exchanges", ())),
    )
    old_groups: dict[tuple[str, str, str, str], list[Mapping[str, Any]]] = defaultdict(
        list
    )
    new_groups: dict[tuple[str, str, str, str], list[Mapping[str, Any]]] = defaultdict(
        list
    )
    for exchange in old_exchanges:
        old_groups[_exchange_group_key(exchange)].append(exchange)
    for exchange in new_exchanges:
        new_groups[_exchange_group_key(exchange)].append(exchange)
    for group_key in sorted(old_groups.keys() | new_groups.keys()):
        pairs, removals, additions = _pair_exchange_group(
            old_groups.get(group_key, []), new_groups.get(group_key, [])
        )
        for old_exchange, new_exchange, exchange_occurrence in pairs:
            exchange_record = _exchange_base(base, new_exchange, exchange_occurrence)
            old_provider = _provider_identity(old_exchange)
            new_provider = _provider_identity(new_exchange)
            if _canonical_json(old_provider) != _canonical_json(new_provider):
                relink = _value_record(
                    exchange_record,
                    object_type="exchange",
                    change_type="supplier relink",
                    changed_field="provider",
                    old=old_provider,
                    new=new_provider,
                    unit=new_exchange.get("unit"),
                )
                relink["old_provider_identity"] = _canonical_json(old_provider)
                relink["new_provider_identity"] = _canonical_json(new_provider)
                records.append(relink)
                if old_exchange.get("location") != new_exchange.get(
                    "location"
                ) and not any(item.proxy for item in attributions):
                    summary.add_fallback(
                        {
                            "scenario": scenario_label,
                            "requested geography": old_exchange.get("location"),
                            "requested technology": old_exchange.get("name"),
                            "selected proxy": new_exchange.get("location"),
                            "fallback rank": None,
                            "explanation": "Supplier location changed during deterministic relinking.",
                            "affected activity count": 1,
                        }
                    )
            old_amount = old_exchange.get("amount")
            new_amount = new_exchange.get("amount")
            if _canonical_json(old_amount) != _canonical_json(new_amount):
                records.append(
                    _value_record(
                        exchange_record,
                        object_type="exchange",
                        change_type="amount change",
                        changed_field="amount",
                        old=old_amount,
                        new=new_amount,
                        unit=new_exchange.get("unit"),
                    )
                )
            for field_name in sorted(UNCERTAINTY_FIELDS):
                old_value = old_exchange.get(field_name)
                new_value = new_exchange.get(field_name)
                if _canonical_json(old_value) == _canonical_json(new_value):
                    continue
                records.append(
                    _value_record(
                        exchange_record,
                        object_type="exchange",
                        change_type="uncertainty change",
                        changed_field=field_name,
                        old=old_value,
                        new=new_value,
                        unit=new_exchange.get("unit"),
                    )
                )
            old_metadata = _exchange_metadata(old_exchange)
            new_metadata = _exchange_metadata(new_exchange)
            for field_name in sorted(old_metadata.keys() | new_metadata.keys()):
                old_value = old_metadata.get(field_name)
                new_value = new_metadata.get(field_name)
                if _canonical_json(old_value) == _canonical_json(new_value):
                    continue
                records.append(
                    _value_record(
                        exchange_record,
                        object_type="exchange",
                        change_type="field modification",
                        changed_field=field_name,
                        old=old_value,
                        new=new_value,
                        unit=new_exchange.get("unit"),
                    )
                )
        for old_exchange, exchange_occurrence in removals:
            exchange_record = _exchange_base(base, old_exchange, exchange_occurrence)
            records.append(
                _value_record(
                    exchange_record,
                    object_type="exchange",
                    change_type="removal",
                    changed_field=None,
                    old=_visible_exchange(old_exchange),
                    new=None,
                    unit=old_exchange.get("unit"),
                )
            )
        for new_exchange, exchange_occurrence in additions:
            exchange_record = _exchange_base(base, new_exchange, exchange_occurrence)
            records.append(
                _value_record(
                    exchange_record,
                    object_type="exchange",
                    change_type="addition",
                    changed_field=None,
                    old=None,
                    new=_visible_exchange(new_exchange),
                    unit=new_exchange.get("unit"),
                )
            )

    if "market" in str(identity_activity.get("name", "")).lower() and any(
        record["object_type"] == "exchange" for record in records
    ):
        _collect_market_row(
            summary, scenario_label, old_activity, new_activity, attributions
        )
    return sorted(
        records,
        key=lambda record: (
            record.get("exchange_type") or "",
            record.get("exchange_name") or "",
            record.get("exchange_product") or "",
            record.get("exchange_location") or "",
            (
                -1
                if record.get("exchange_occurrence") is None
                else record["exchange_occurrence"]
            ),
            record.get("changed_field") or "",
            record["change_type"],
        ),
    )


def _collect_market_row(
    summary: _Summary,
    scenario: str,
    old_activity: Mapping[str, Any],
    new_activity: Mapping[str, Any],
    attributions: Sequence[Attribution],
) -> None:
    def suppliers(activity):
        return [
            exchange
            for exchange in activity.get("exchanges", ())
            if exchange.get("type") == "technosphere"
        ]

    old_suppliers = suppliers(old_activity)
    new_suppliers = suppliers(new_activity)
    old_shares = {
        _canonical_json(_provider_identity(exchange)): _number(exchange.get("amount"))
        or 0.0
        for exchange in old_suppliers
    }
    new_shares = {
        _canonical_json(_provider_identity(exchange)): _number(exchange.get("amount"))
        or 0.0
        for exchange in new_suppliers
    }
    movements = [
        (abs(new_shares.get(key, 0.0) - old_shares.get(key, 0.0)), key)
        for key in old_shares.keys() | new_shares.keys()
    ]
    largest = max(movements, default=(0.0, None))
    algorithm = next((item.algorithm for item in attributions if item.algorithm), None)
    summary.market_rows.append(
        {
            "scenario": scenario,
            "activity": old_activity.get("name"),
            "location": old_activity.get("location"),
            "algorithm": algorithm,
            "old supplier count": len(old_suppliers),
            "new supplier count": len(new_suppliers),
            "old share total": sum(old_shares.values()),
            "new share total": sum(new_shares.values()),
            "largest share movement": largest[0],
            "largest movement supplier": largest[1],
            "missing technologies": ", ".join(
                sorted(set(old_shares) - set(new_shares))
            ),
            "dropped share": sum(
                old_shares[key] for key in set(old_shares) - set(new_shares)
            ),
            "market basis": (
                "marginal" if algorithm and "marginal" in algorithm else "average"
            ),
        }
    )


def _summary_category(record: Mapping[str, Any]) -> str:
    object_type = record["object_type"]
    change = record["change_type"]
    if change == "addition":
        action = "added"
    elif change == "removal":
        action = "removed"
    elif change == "supplier relink":
        action = "relinked"
    else:
        action = "modified"
    return f"{object_type} {action}"


def _object_token(record: Mapping[str, Any], category: str) -> tuple[Any, ...]:
    activity = (
        record.get("activity_code"),
        record.get("activity_name"),
        record.get("activity_product"),
        record.get("activity_location"),
        record.get("activity_occurrence"),
    )
    if record["object_type"] == "activity":
        return activity
    return activity + (
        record.get("exchange_type"),
        record.get("exchange_name"),
        record.get("exchange_product"),
        record.get("exchange_location"),
        record.get("exchange_occurrence"),
        "relink" if category.endswith("relinked") else "exchange",
    )


class _ParquetSink:
    def __init__(self, path: Path) -> None:
        self.path = path
        self.writer = pq.ParquetWriter(
            path,
            DETAIL_SCHEMA,
            compression="zstd",
            use_dictionary=True,
            write_statistics=True,
        )
        self.rows: list[dict[str, Any]] = []

    def write(self, rows: Iterable[Mapping[str, Any]]) -> None:
        for row in rows:
            self.rows.append(
                {field.name: row.get(field.name) for field in DETAIL_SCHEMA}
            )
            if len(self.rows) >= DETAIL_BATCH_SIZE:
                self.flush()

    def flush(self) -> None:
        if not self.rows:
            return
        table = pa.Table.from_pylist(self.rows, schema=DETAIL_SCHEMA)
        self.writer.write_table(table, row_group_size=DETAIL_BATCH_SIZE)
        self.rows.clear()

    def close(self) -> None:
        self.flush()
        self.writer.close()


def _certificate_key(report: ValidationReport | None) -> str | None:
    return None if report is None else report.certificate_key


def _cache_key(
    source_fingerprint: str, scenarios: Sequence[ReportScenario]
) -> tuple[Any, ...]:
    return (
        REPORT_SCHEMA_VERSION,
        source_fingerprint,
        tuple(
            (
                scenario.identity,
                scenario.store.generation,
                _certificate_key(scenario.validation_report),
            )
            for scenario in scenarios
        ),
    )


def _new_report_id(build_id: str) -> str:
    timestamp = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%S%fZ")
    return f"{timestamp}-{build_id[:8]}"


def _resolve_paths(
    filepath: str | Path | None,
    name: str | None,
    report_id: str,
) -> tuple[Path, Path]:
    if filepath is None:
        directory = Path.cwd() / "export" / "change reports"
        workbook_name = name or f"change-report-{report_id}.xlsx"
        workbook_path = directory / workbook_name
    else:
        path = Path(filepath).expanduser()
        if path.suffix.lower() == ".xlsx":
            if name is not None:
                raise ValueError("name cannot be combined with an .xlsx filepath.")
            workbook_path = path
        else:
            workbook_name = name or f"change-report-{report_id}.xlsx"
            workbook_path = path / workbook_name
    if workbook_path.suffix.lower() != ".xlsx":
        workbook_path = workbook_path.with_suffix(".xlsx")
    if report_id not in workbook_path.stem:
        workbook_path = workbook_path.with_name(
            f"{workbook_path.stem}-{report_id}{workbook_path.suffix}"
        )
    workbook_path.parent.mkdir(parents=True, exist_ok=True)
    if workbook_path.exists():
        workbook_path = workbook_path.with_name(
            f"{workbook_path.stem}-{uuid.uuid4().hex[:6]}{workbook_path.suffix}"
        )
    return workbook_path, workbook_path.with_suffix(".parquet")


def _collect_event_summaries(
    summary: _Summary, scenario: ReportScenario, scenario_label: str
) -> None:
    payload = scenario.provenance_payload or {}
    fallback_counts: Counter = Counter()
    fallback_payloads = {}
    methodology: dict[tuple[Any, ...], tuple[str, str | None]] = {}
    for item in payload.get("events", ()) if isinstance(payload, Mapping) else ():
        try:
            event = ProvenanceEvent.from_dict(item)
        except (KeyError, TypeError, ValueError):
            continue
        if event.proxy is not None or event.fallback_rank is not None:
            key = (
                event.activity.get("location"),
                event.activity.get("name"),
                event.proxy,
                event.fallback_rank,
                event.explanation,
            )
            fallback_counts[key] += 1
            fallback_payloads[key] = event
        computed_target_values = None
        if event.reason_code.startswith(
            "metals.material_rule"
        ) or event.reason_code == ("metals.post_allocation_resource_correction"):
            computed_target_values = _canonical_json(event.computed_target_values)
        methodology.setdefault(
            (
                event.sector,
                event.transformation,
                event.algorithm,
                event.iam_variable,
                event.configuration_reference,
                event.reason_code,
            ),
            (event.explanation, computed_target_values),
        )
    for key, count in sorted(fallback_counts.items(), key=lambda item: str(item[0])):
        summary.add_fallback(
            {
                "scenario": scenario_label,
                "requested geography": key[0],
                "requested technology": key[1],
                "selected proxy": key[2],
                "fallback rank": key[3],
                "explanation": key[4],
                "affected activity count": count,
            }
        )
    for row, values in sorted(methodology.items(), key=lambda item: str(item[0])):
        explanation, computed_target_values = values
        summary.methodology_rows.append(
            {
                "scenario": scenario_label,
                "sector": row[0],
                "transformation": row[1],
                "algorithm": row[2],
                "IAM variable": row[3],
                "configuration reference": row[4],
                "reason code": row[5],
                "explanation": explanation,
                "computed target values": computed_target_values,
            }
        )


def _generate_details(
    *,
    source_store: InventoryStore,
    scenarios: Sequence[ReportScenario],
    details_path: Path,
    report_id: str,
    build_id: str,
    source_fingerprint: str,
) -> _Summary:
    source_index, source_exchange_count, source_visible_fingerprint = _activity_index(
        source_store
    )
    effective_source_fingerprint = source_fingerprint or source_visible_fingerprint
    summary = _Summary(
        source_activity_count=len(source_index),
        source_exchange_count=source_exchange_count,
    )
    temporary = details_path.with_name(f".{details_path.name}.tmp-{uuid.uuid4().hex}")
    sink = _ParquetSink(temporary)
    try:
        for scenario_order, scenario in enumerate(scenarios):
            scenario_label = _scenario_label(scenario.identity)
            final_index, final_exchange_count, final_fingerprint = _activity_index(
                scenario.store
            )
            summary.final_counts[scenario_label] = (
                len(final_index),
                final_exchange_count,
            )
            summary.final_fingerprints[scenario_label] = final_fingerprint
            _collect_event_summaries(summary, scenario, scenario_label)
            attribution_index = _AttributionIndex(scenario)
            for pair in _pair_activities(source_index, final_index):
                records = _activity_records(
                    source_store,
                    scenario.store,
                    pair,
                    report_id=report_id,
                    build_id=build_id,
                    scenario_order=scenario_order,
                    scenario_label=scenario_label,
                    source_fingerprint=effective_source_fingerprint,
                    final_fingerprint=final_fingerprint,
                    certificate_key=_certificate_key(scenario.validation_report),
                    attribution_index=attribution_index,
                    summary=summary,
                )
                summary.consume_activity(records)
                sink.write(records)
        sink.close()
        os.replace(temporary, details_path)
    except Exception:
        try:
            sink.close()
        finally:
            temporary.unlink(missing_ok=True)
        raise
    return summary


def _validation_rows(
    scenarios: Sequence[ReportScenario],
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    findings = []
    coverage = []
    for scenario in scenarios:
        report = scenario.validation_report
        if report is None:
            continue
        scenario_label = _scenario_label(scenario.identity)
        for phase in report.phase_results:
            for result in phase.rule_results:
                successful = not any(not issue.suppressed for issue in result.issues)
                suppressed = [issue for issue in result.issues if issue.suppressed]
                coverage.append(
                    {
                        "scenario": scenario_label,
                        "phase": phase.phase_id,
                        "rule ID": result.rule_id,
                        "applicability": result.applicability,
                        "checked objects": result.checked_object_count,
                        "successful": successful,
                        "suppressed count": len(suppressed),
                        "suppression explanations": " | ".join(
                            dict.fromkeys(
                                issue.suppression_explanation or ""
                                for issue in suppressed
                            )
                        ),
                    }
                )
                for issue in result.issues:
                    if issue.suppressed:
                        continue
                    findings.append(
                        {
                            "scenario": scenario_label,
                            "severity": issue.severity,
                            "rule ID": issue.rule_id,
                            "phase": phase.phase_id,
                            "message": issue.message,
                            "activity": (
                                _canonical_json(issue.activity_key)
                                if issue.activity_key is not None
                                else None
                            ),
                            "activity code": issue.activity_code,
                            "exchange": issue.exchange_id,
                            "expected": (
                                _canonical_json(issue.expected)
                                if issue.expected is not None
                                else None
                            ),
                            "actual": (
                                _canonical_json(issue.actual)
                                if issue.actual is not None
                                else None
                            ),
                            "tolerance": issue.tolerance,
                        }
                    )
    return findings, coverage


HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
ERROR_FILL = PatternFill("solid", fgColor="F4CCCC")
WARNING_FILL = PatternFill("solid", fgColor="FFF2CC")
PASSED_FILL = PatternFill("solid", fgColor="D9EAD3")


def _write_table_sheet(
    workbook: Workbook,
    title: str,
    rows: Sequence[Mapping[str, Any]],
    headers: Sequence[str],
    table_index: int,
) -> openpyxl.worksheet.worksheet.Worksheet:
    worksheet = workbook.create_sheet(title)
    worksheet.append(list(headers))
    if rows:
        for row in rows:
            worksheet.append([row.get(header) for header in headers])
    else:
        worksheet.append([None] * len(headers))
    for cell in worksheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    worksheet.freeze_panes = "A2"
    table = Table(
        displayName=f"ReportTable{table_index}",
        ref=f"A1:{get_column_letter(len(headers))}{worksheet.max_row}",
    )
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    worksheet.add_table(table)
    for column, header in enumerate(headers, 1):
        maximum = max(
            len(str(header)),
            *(
                len(str(worksheet.cell(row=row, column=column).value or ""))
                for row in range(2, min(worksheet.max_row, 200) + 1)
            ),
        )
        worksheet.column_dimensions[get_column_letter(column)].width = min(
            max(maximum + 2, 10), 55
        )
        for row in range(2, worksheet.max_row + 1):
            cell = worksheet.cell(row=row, column=column)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if any(
                token in header.lower()
                for token in ("delta", "share", "tolerance", "value")
            ) and isinstance(cell.value, (int, float)):
                cell.number_format = "0.000000E+00"
    return worksheet


def _summary_rows(summary: _Summary) -> list[dict[str, Any]]:
    categories = (
        "activity added",
        "activity removed",
        "activity modified",
        "activity relinked",
        "exchange added",
        "exchange removed",
        "exchange modified",
        "exchange relinked",
    )
    rows = []
    for scenario in sorted(summary.scenario_counts):
        row = {"scenario": scenario}
        row.update(
            {
                category: summary.scenario_counts[scenario][category]
                for category in categories
            }
        )
        final_activities, final_exchanges = summary.final_counts[scenario]
        row.update(
            {
                "activities before": summary.source_activity_count,
                "activities after": final_activities,
                "exchanges before": summary.source_exchange_count,
                "exchanges after": final_exchanges,
            }
        )
        rows.append(row)
    return rows


def _sector_rows(summary: _Summary) -> list[dict[str, Any]]:
    rows = []
    for (scenario, sector, transformation), counts in sorted(
        summary.sector_counts.items()
    ):
        row = {
            "scenario": scenario,
            "sector": sector,
            "transformation": transformation,
        }
        for object_type in ("activity", "exchange"):
            for action in ("added", "removed", "modified", "relinked"):
                row[f"{object_type} {action}"] = counts[f"{object_type} {action}"]
        rows.append(row)
    return rows


def _write_overview(
    workbook: Workbook,
    *,
    artifacts: ChangeReportArtifacts,
    scenarios: Sequence[ReportScenario],
    summary: _Summary,
    source_database: str | None,
    source_type: str | None,
    version: str | None,
    system_model: str | None,
    premise_version: str,
    details_path: Path | None,
) -> None:
    worksheet = workbook.active
    worksheet.title = "Overview"
    unattributed_count = sum(
        sum(counts.values())
        for (_, sector, transformation), counts in summary.sector_counts.items()
        if sector == "unattributed" or transformation == "unattributed"
    )
    rows = [
        ("Report ID", artifacts.report_id),
        ("Status", artifacts.status),
        ("Generated (UTC)", datetime.now(timezone.utc).isoformat()),
        ("Source database", source_database),
        ("Source format", source_type),
        ("Ecoinvent version", version),
        ("System model", system_model),
        ("premise version", premise_version),
        ("Report schema", REPORT_SCHEMA_VERSION),
        ("Provenance schema", PROVENANCE_SCHEMA_VERSION),
        ("Source fingerprint", artifacts.source_fingerprint),
        (
            "Validation ruleset",
            next(
                (
                    s.validation_report.ruleset_version
                    for s in scenarios
                    if s.validation_report
                ),
                None,
            ),
        ),
        ("Scenario count", len(scenarios)),
        ("Source activities", summary.source_activity_count),
        ("Source exchanges", summary.source_exchange_count),
        ("Unattributed change objects", unattributed_count),
        ("Detailed audit", None if details_path is None else details_path.name),
    ]
    for position, scenario in enumerate(scenarios, 1):
        label = _scenario_label(scenario.identity)
        rows.extend(
            [
                (f"Scenario {position} identity", label),
                (
                    f"Scenario {position} final fingerprint",
                    summary.final_fingerprints.get(label),
                ),
                (
                    f"Scenario {position} validation status",
                    (
                        "not available"
                        if scenario.validation_report is None
                        else (
                            "passed" if scenario.validation_report.valid else "failed"
                        )
                    ),
                ),
                (
                    f"Scenario {position} certificate key",
                    _certificate_key(scenario.validation_report),
                ),
            ]
        )
    worksheet.append(["Field", "Value"])
    for row in rows:
        worksheet.append(row)
    for cell in worksheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    status_cell = worksheet[3][1]
    status_cell.fill = PASSED_FILL if artifacts.status == "passed" else ERROR_FILL
    if unattributed_count:
        unattributed_row = next(
            row
            for row in range(2, worksheet.max_row + 1)
            if worksheet.cell(row, 1).value == "Unattributed change objects"
        )
        worksheet.cell(row=unattributed_row, column=2).fill = WARNING_FILL
    if details_path is not None:
        audit_row = next(
            row
            for row in range(2, worksheet.max_row + 1)
            if worksheet.cell(row, 1).value == "Detailed audit"
        )
        audit_cell = worksheet.cell(row=audit_row, column=2)
        audit_cell.hyperlink = os.path.relpath(
            details_path, artifacts.workbook_path.parent
        )
        audit_cell.style = "Hyperlink"
    start = len(rows) + 4
    worksheet.cell(start, 1, "How to read this report")
    worksheet.cell(start, 1).font = Font(bold=True, size=13)
    definitions = [
        (
            "Addition / removal",
            "An object exists on only one side of the certified source-to-final comparison.",
        ),
        (
            "Modification",
            "A user-visible field, amount, uncertainty parameter, or nested metadata value changed.",
        ),
        (
            "Relink",
            "A matched technosphere exchange points to a different provider identity or geography.",
        ),
        (
            "Percentage change",
            "(new - old) / abs(old); zero baselines are ranked by absolute delta.",
        ),
        (
            "Suppression",
            "A documented validation exception; suppressed findings are counted in Validation Coverage.",
        ),
        (
            "Unattributed",
            "No structured provenance event or certified transformation scope matched the difference.",
        ),
    ]
    for offset, definition in enumerate(definitions, 1):
        worksheet.cell(start + offset, 1, definition[0]).font = Font(bold=True)
        worksheet.cell(start + offset, 2, definition[1])
    worksheet.column_dimensions["A"].width = 28
    worksheet.column_dimensions["B"].width = 90
    for row in worksheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    worksheet.freeze_panes = "A2"


def _write_workbook(
    *,
    workbook_path: Path,
    details_path: Path | None,
    artifacts: ChangeReportArtifacts,
    scenarios: Sequence[ReportScenario],
    summary: _Summary,
    source_database: str | None,
    source_type: str | None,
    version: str | None,
    system_model: str | None,
    premise_version: str,
) -> None:
    workbook = Workbook()
    _write_overview(
        workbook,
        artifacts=artifacts,
        scenarios=scenarios,
        summary=summary,
        source_database=source_database,
        source_type=source_type,
        version=version,
        system_model=system_model,
        premise_version=premise_version,
        details_path=details_path,
    )
    scenario_rows = _summary_rows(summary)
    scenario_headers = [
        "scenario",
        "activities before",
        "activities after",
        "exchanges before",
        "exchanges after",
        "activity added",
        "activity removed",
        "activity modified",
        "activity relinked",
        "exchange added",
        "exchange removed",
        "exchange modified",
        "exchange relinked",
    ]
    worksheet = _write_table_sheet(
        workbook, "Scenario Summary", scenario_rows, scenario_headers, 1
    )
    for row in range(2, worksheet.max_row + 1):
        for column in range(6, len(scenario_headers) + 1):
            worksheet.cell(row, column).hyperlink = "#'Sector Summary'!A1"
            worksheet.cell(row, column).style = "Hyperlink"
    sector_headers = [
        "scenario",
        "sector",
        "transformation",
        "activity added",
        "activity removed",
        "activity modified",
        "activity relinked",
        "exchange added",
        "exchange removed",
        "exchange modified",
        "exchange relinked",
    ]
    sector_sheet = _write_table_sheet(
        workbook, "Sector Summary", _sector_rows(summary), sector_headers, 2
    )
    for row in range(2, sector_sheet.max_row + 1):
        for column in range(4, len(sector_headers) + 1):
            sector_sheet.cell(row, column).hyperlink = "#'Key Changes'!A1"
            sector_sheet.cell(row, column).style = "Hyperlink"
    key_rows = summary.key_change_rows()
    key_headers = (
        list(key_rows[0])
        if key_rows
        else [
            "scenario",
            "sector",
            "transformation",
            "object type",
            "change type",
            "activity",
            "product",
            "location",
            "exchange",
            "field",
            "old value",
            "new value",
            "absolute delta",
            "relative delta",
            "unit",
            "reason",
        ]
    )
    _write_table_sheet(workbook, "Key Changes", key_rows, key_headers, 3)
    market_headers = [
        "scenario",
        "activity",
        "location",
        "algorithm",
        "old supplier count",
        "new supplier count",
        "old share total",
        "new share total",
        "largest share movement",
        "largest movement supplier",
        "missing technologies",
        "dropped share",
        "market basis",
    ]
    _write_table_sheet(
        workbook, "Market Changes", summary.market_rows, market_headers, 4
    )
    fallback_headers = [
        "scenario",
        "requested geography",
        "requested technology",
        "selected proxy",
        "fallback rank",
        "explanation",
        "affected activity count",
    ]
    _write_table_sheet(
        workbook, "Fallbacks & Proxies", summary.fallback_rows(), fallback_headers, 5
    )
    findings, coverage = _validation_rows(scenarios)
    finding_headers = [
        "scenario",
        "severity",
        "rule ID",
        "phase",
        "message",
        "activity",
        "activity code",
        "exchange",
        "expected",
        "actual",
        "tolerance",
    ]
    finding_sheet = _write_table_sheet(
        workbook, "Validation Findings", findings, finding_headers, 6
    )
    for row in range(2, finding_sheet.max_row + 1):
        severity = finding_sheet.cell(row, 2).value
        if severity == "error":
            for cell in finding_sheet[row]:
                cell.fill = ERROR_FILL
        elif severity == "warning":
            for cell in finding_sheet[row]:
                cell.fill = WARNING_FILL
    coverage_headers = [
        "scenario",
        "phase",
        "rule ID",
        "applicability",
        "checked objects",
        "successful",
        "suppressed count",
        "suppression explanations",
    ]
    _write_table_sheet(workbook, "Validation Coverage", coverage, coverage_headers, 7)
    methodology_rows = list(summary.methodology_rows)
    methodology_rows.append(
        {
            "scenario": "source",
            "sector": "source",
            "transformation": "normalization baseline",
            "algorithm": None,
            "IAM variable": None,
            "configuration reference": None,
            "reason code": artifacts.source_fingerprint,
            "explanation": "Fingerprint of the normalized source inventory used by this report.",
        }
    )
    for scenario in scenarios:
        report = scenario.validation_report
        underlying = scenario.store
        while hasattr(underlying, "_store"):
            underlying = underlying._store
        certificate = getattr(underlying, "_validation_certificate_payload", None)
        methodology_rows.append(
            {
                "scenario": _scenario_label(scenario.identity),
                "sector": "validation",
                "transformation": "certification",
                "algorithm": None,
                "IAM variable": None,
                "configuration reference": None,
                "reason code": _certificate_key(report),
                "explanation": "Immutable validation certificate key for the reported store generation.",
            }
        )
        if isinstance(certificate, Mapping):
            methodology_rows.append(
                {
                    "scenario": _scenario_label(scenario.identity),
                    "sector": "IAM",
                    "transformation": "scenario input",
                    "algorithm": None,
                    "IAM variable": None,
                    "configuration reference": None,
                    "reason code": certificate.get("iam_fingerprint"),
                    "explanation": "Fingerprint of the IAM inputs used for certification.",
                }
            )
    methodology_headers = [
        "scenario",
        "sector",
        "transformation",
        "algorithm",
        "IAM variable",
        "configuration reference",
        "reason code",
        "explanation",
        "computed target values",
    ]
    _write_table_sheet(
        workbook, "Methodology", methodology_rows, methodology_headers, 8
    )
    workbook.properties.title = "premise structured change report"
    workbook.properties.subject = f"Report schema {REPORT_SCHEMA_VERSION}"
    workbook.properties.creator = "premise"
    workbook.calculation.fullCalcOnLoad = False
    workbook.save(workbook_path)


_REPORT_LOCK = threading.RLock()


def generate_structured_change_report(
    *,
    source_store: InventoryStore,
    scenarios: Sequence[ReportScenario],
    build_id: str,
    source_fingerprint: str,
    status: Literal["passed", "failed"] = "passed",
    filepath: str | Path | None = None,
    name: str | None = None,
    source_database: str | None = None,
    source_type: str | None = None,
    version: str | None = None,
    system_model: str | None = None,
    premise_version: str = "unknown",
    cache_entry: ChangeReportCacheEntry | None = None,
) -> GeneratedChangeReport:
    """Generate or refresh a V2 workbook and its cached detailed audit."""

    if status not in {"passed", "failed"}:
        raise ValueError("status must be either 'passed' or 'failed'.")
    if not scenarios:
        raise RuntimeError("A change report requires at least one updated scenario.")
    key = _cache_key(source_fingerprint, scenarios)
    with _REPORT_LOCK:
        can_reuse = (
            cache_entry is not None
            and cache_entry.cache_key == key
            and cache_entry.details_path.is_file()
        )
        report_id = cache_entry.report_id if can_reuse else _new_report_id(build_id)
        workbook_path, proposed_details_path = _resolve_paths(filepath, name, report_id)
        if can_reuse:
            details_path = cache_entry.details_path
            summary = cache_entry.summary
            effective_cache = cache_entry
        else:
            details_path = proposed_details_path
            if details_path.exists():
                details_path = details_path.with_name(
                    f"{details_path.stem}-{uuid.uuid4().hex[:6]}{details_path.suffix}"
                )
            summary = _generate_details(
                source_store=source_store,
                scenarios=scenarios,
                details_path=details_path,
                report_id=report_id,
                build_id=build_id,
                source_fingerprint=source_fingerprint,
            )
            effective_cache = ChangeReportCacheEntry(
                cache_key=key,
                report_id=report_id,
                details_path=details_path,
                summary=summary,
            )
        artifacts = ChangeReportArtifacts(
            report_id=report_id,
            status=status,
            workbook_path=workbook_path.resolve(),
            details_path=details_path.resolve(),
            scenario_identities=tuple(scenario.identity for scenario in scenarios),
            source_fingerprint=source_fingerprint,
            validation_certificate_keys=tuple(
                key
                for key in (
                    _certificate_key(scenario.validation_report)
                    for scenario in scenarios
                )
                if key is not None
            ),
        )
        _write_workbook(
            workbook_path=workbook_path,
            details_path=details_path,
            artifacts=artifacts,
            scenarios=scenarios,
            summary=summary,
            source_database=source_database,
            source_type=source_type,
            version=version,
            system_model=system_model,
            premise_version=premise_version,
        )
        return GeneratedChangeReport(artifacts=artifacts, cache_entry=effective_cache)


def generate_validation_diagnostic_workbook(
    *,
    scenarios: Sequence[ReportScenario],
    build_id: str,
    source_fingerprint: str,
    filepath: str | Path | None = None,
    name: str | None = None,
    source_database: str | None = None,
    source_type: str | None = None,
    version: str | None = None,
    system_model: str | None = None,
    premise_version: str = "unknown",
) -> ChangeReportArtifacts:
    """Write a validation-only workbook when an exhaustive failed diff is unsafe."""

    report_id = _new_report_id(build_id)
    workbook_path, _ = _resolve_paths(filepath, name, report_id)
    summary = _Summary()
    for scenario in scenarios:
        label = _scenario_label(scenario.identity)
        try:
            activity_count = len(scenario.store)
        except Exception:
            activity_count = 0
        summary.final_counts[label] = (activity_count, 0)
        summary.final_fingerprints[label] = "unavailable"
        _collect_event_summaries(summary, scenario, label)
    artifacts = ChangeReportArtifacts(
        report_id=report_id,
        status="failed",
        workbook_path=workbook_path.resolve(),
        details_path=None,
        scenario_identities=tuple(scenario.identity for scenario in scenarios),
        source_fingerprint=source_fingerprint,
        validation_certificate_keys=tuple(
            key
            for key in (
                _certificate_key(scenario.validation_report) for scenario in scenarios
            )
            if key is not None
        ),
    )
    _write_workbook(
        workbook_path=workbook_path,
        details_path=None,
        artifacts=artifacts,
        scenarios=scenarios,
        summary=summary,
        source_database=source_database,
        source_type=source_type,
        version=version,
        system_model=system_model,
        premise_version=premise_version,
    )
    return artifacts


__all__ = [
    "ChangeReportArtifacts",
    "ChangeReportCacheEntry",
    "GeneratedChangeReport",
    "REPORT_SCHEMA_VERSION",
    "ReportScenario",
    "generate_structured_change_report",
    "generate_validation_diagnostic_workbook",
]
