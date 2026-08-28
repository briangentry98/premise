from __future__ import annotations

import json
from pathlib import Path

import openpyxl
import pyarrow.parquet as pq
import pytest
import premise

from premise.change_report import (
    DETAIL_SCHEMA,
    REPORT_SCHEMA_VERSION,
    ReportScenario,
    generate_structured_change_report,
)
from premise.inventory_store import CompactInventoryStore
from premise.new_database import NewDatabase
from premise.provenance import ProvenanceCollector, record_change_event
from premise.validation_framework import (
    PremiseValidationError,
    ValidationIssue,
    ValidationPhaseResult,
    ValidationReport,
    ValidationRuleResult,
    inventory_store_fingerprint,
)


def _activity(code, name, location, exchanges, **extra):
    return {
        "database": "source-db",
        "code": code,
        "name": name,
        "reference product": f"product {name}",
        "location": location,
        "unit": "kilogram",
        "exchanges": exchanges,
        **extra,
    }


def _production(code, name, location):
    return {
        "type": "production",
        "name": name,
        "product": f"product {name}",
        "location": location,
        "unit": "kilogram",
        "amount": 1.0,
        "input": ("source-db", code),
    }


def _synthetic_stores():
    source = [
        _activity(
            "a",
            "consumer",
            "CH",
            [
                _production("a", "consumer", "CH"),
                {
                    "type": "technosphere",
                    "name": "supplier",
                    "product": "product supplier",
                    "location": "CH",
                    "unit": "kilogram",
                    "amount": 1.0,
                    "uncertainty type": 5,
                    "minimum": 0.5,
                    "maximum": 1.5,
                    "input": ("source-db", "b"),
                    "comment": "source | value\nwith newline",
                },
            ],
            comment="source comment",
            _runtime_cache="ignored",
        ),
        _activity(
            "b",
            "supplier",
            "CH",
            [_production("b", "supplier", "CH")],
        ),
        _activity(
            None,
            "semantic duplicate",
            "GLO",
            [_production("missing", "semantic duplicate", "GLO")],
            comment="first",
        ),
        _activity(
            None,
            "semantic duplicate",
            "GLO",
            [_production("missing", "semantic duplicate", "GLO")],
            comment="second",
        ),
    ]
    final = [
        _activity(
            "a",
            "consumer",
            "CH",
            [
                {**_production("a", "consumer", "CH"), "input": ("export-db", "a")},
                {
                    "type": "technosphere",
                    "name": "supplier",
                    "product": "product supplier",
                    "location": "RER",
                    "unit": "kilogram",
                    "amount": 2.0,
                    "uncertainty type": 5,
                    "minimum": 0.75,
                    "maximum": 2.5,
                    "input": ("export-db", "b"),
                    "comment": "final | value\nwith newline",
                },
            ],
            database="export-db",
            comment="final comment",
            _runtime_cache="different but ignored",
        ),
        _activity(
            "b",
            "supplier",
            "RER",
            [_production("b", "supplier", "RER")],
            database="export-db",
        ),
        _activity(
            None,
            "semantic duplicate",
            "GLO",
            [_production("missing", "semantic duplicate", "GLO")],
            comment="second changed",
        ),
        _activity(
            None,
            "semantic duplicate",
            "GLO",
            [_production("missing", "semantic duplicate", "GLO")],
            comment="first",
        ),
        _activity(
            "c",
            "added",
            "GLO",
            [_production("c", "added", "GLO")],
        ),
    ]
    return CompactInventoryStore(source), CompactInventoryStore(final)


def test_structured_change_report_exact_diff_and_workbook(tmp_path):
    source, final = _synthetic_stores()
    collector = ProvenanceCollector("build-id")
    with collector.session(("image", "path", 2050, ()), "electricity"):
        record_change_event(
            type("Transformation", (), {"system_model": "cutoff"})(),
            final.activity(0).to_dict(),
            "updated",
            sector="electricity",
        )

    generated = generate_structured_change_report(
        source_store=source,
        scenarios=(
            ReportScenario(
                identity=("image", "path", 2050, ()),
                store=final,
                provenance_payload=collector.payload_for(("image", "path", 2050, ())),
            ),
        ),
        build_id="1234567890abcdef",
        source_fingerprint="source-fingerprint",
        filepath=tmp_path,
        name="audit.xlsx",
        source_database="source-db",
        source_type="brightway",
        version="3.12",
        system_model="cutoff",
        premise_version="3.0.0",
    )

    artifacts = generated.artifacts
    assert artifacts.status == "passed"
    assert artifacts.workbook_path.is_file()
    assert artifacts.details_path.is_file()

    table = pq.read_table(artifacts.details_path)
    assert table.schema.names == DETAIL_SCHEMA.names
    assert table.schema.remove_metadata() == DETAIL_SCHEMA.remove_metadata()
    assert table.schema.metadata[b"premise_report_schema_version"] == b"2"
    rows = table.to_pylist()
    assert rows
    assert all(row["report_schema_version"] == REPORT_SCHEMA_VERSION for row in rows)
    assert any(row["change_type"] == "supplier relink" for row in rows)
    assert any(row["change_type"] == "amount change" for row in rows)
    assert any(row["change_type"] == "uncertainty change" for row in rows)
    assert any(row["change_type"] == "addition" for row in rows)
    assert any(
        row["changed_field"] == "comment"
        and row["old_value_json"] is not None
        and json.loads(row["old_value_json"]) == "source | value\nwith newline"
        for row in rows
    )
    assert not any(
        row["changed_field"] in {"database", "_runtime_cache"} for row in rows
    )
    assert any(row["transformations"] == ["electricity"] for row in rows)
    assert any(row["transformations"] == ["unattributed"] for row in rows)

    order = [
        (
            row["scenario_order"],
            row["activity_name"] or "",
            row["activity_product"] or "",
            row["activity_location"] or "",
            row["activity_occurrence"],
            row["exchange_type"] or "",
            row["exchange_name"] or "",
            (
                row["exchange_occurrence"]
                if row["exchange_occurrence"] is not None
                else -1
            ),
            row["changed_field"] or "",
            row["change_type"],
        )
        for row in rows
    ]
    assert order == sorted(order)

    workbook = openpyxl.load_workbook(artifacts.workbook_path, read_only=False)
    assert workbook.sheetnames == [
        "Overview",
        "Scenario Summary",
        "Sector Summary",
        "Key Changes",
        "Market Changes",
        "Fallbacks & Proxies",
        "Validation Findings",
        "Validation Coverage",
        "Methodology",
    ]
    assert workbook["Scenario Summary"].freeze_panes == "A2"
    overview = workbook["Overview"]
    audit_row = next(
        row
        for row in range(1, overview.max_row + 1)
        if overview.cell(row, 1).value == "Detailed audit"
    )
    assert overview.cell(audit_row, 2).hyperlink is not None
    assert len(workbook["Scenario Summary"].tables) == 1


def test_details_cache_is_reused_but_workbook_is_refreshed(tmp_path):
    source, final = _synthetic_stores()
    kwargs = dict(
        source_store=source,
        scenarios=(ReportScenario(identity=("image", "path", 2050, ()), store=final),),
        build_id="1234567890abcdef",
        source_fingerprint="source-fingerprint",
        filepath=tmp_path,
        premise_version="3.0.0",
    )
    first = generate_structured_change_report(**kwargs)
    first_bytes = first.artifacts.details_path.read_bytes()
    second = generate_structured_change_report(
        **kwargs, cache_entry=first.cache_entry, status="failed"
    )
    assert second.artifacts.details_path == first.artifacts.details_path
    assert second.artifacts.details_path.read_bytes() == first_bytes
    assert second.artifacts.workbook_path != first.artifacts.workbook_path
    assert second.artifacts.status == "failed"


def test_detail_content_is_stable_excluding_report_and_build_identity(tmp_path):
    source, final = _synthetic_stores()
    scenario = (ReportScenario(identity=("image", "path", 2050, ()), store=final),)
    first = generate_structured_change_report(
        source_store=source,
        scenarios=scenario,
        build_id="first-build-id",
        source_fingerprint="source-fingerprint",
        filepath=tmp_path / "first",
    )
    second = generate_structured_change_report(
        source_store=source,
        scenarios=scenario,
        build_id="second-build-id",
        source_fingerprint="source-fingerprint",
        filepath=tmp_path / "second",
    )

    def normalized_rows(path):
        rows = pq.read_table(path).to_pylist()
        for row in rows:
            row.pop("report_id")
            row.pop("build_id")
        return rows

    assert normalized_rows(first.artifacts.details_path) == normalized_rows(
        second.artifacts.details_path
    )


def test_generate_change_report_before_update_is_clear_runtime_error():
    database = object.__new__(NewDatabase)
    database.scenarios = [{"model": "image", "pathway": "path", "year": 2050}]
    with pytest.raises(RuntimeError, match="Call update"):
        database.generate_change_report()


def test_explicit_report_works_when_automatic_reports_are_disabled(tmp_path):
    source, final = _synthetic_stores()
    database = object.__new__(NewDatabase)
    database.scenarios = [
        {
            "model": "image",
            "pathway": "path",
            "year": 2050,
            "applied functions": ["electricity"],
            "_inventory_store": final,
        }
    ]
    database._source_inventory_store = source
    database._compact_source_checkpoint = None
    database._validation_enabled = False
    database._validation_reports = {}
    database._validation_iam_fingerprints = {}
    database.generate_reports = False
    database.build_id = "1234567890abcdef"
    database.source = "source-db"
    database.source_type = "brightway"
    database.version = "3.12"
    database.system_model = "cutoff"
    before = (
        inventory_store_fingerprint(source),
        inventory_store_fingerprint(final),
    )

    artifacts = database.generate_change_report(filepath=tmp_path)

    assert isinstance(artifacts, premise.ChangeReportArtifacts)
    assert artifacts.workbook_path.is_file()
    assert before == (
        inventory_store_fingerprint(source),
        inventory_store_fingerprint(final),
    )


def test_provenance_round_trips_with_inventory_checkpoint(tmp_path):
    _, store = _synthetic_stores()
    collector = ProvenanceCollector("build-id")
    with collector.session(("image", "path", 2050, ()), "electricity"):
        record_change_event(object(), store.activity(0).to_dict(), "updated")
    payload = collector.payload_for(("image", "path", 2050, ()))
    store._provenance_payload = payload

    checkpoint = store.checkpoint(tmp_path / "scenario.inventory-store")
    reopened = CompactInventoryStore.open(checkpoint)

    assert reopened._provenance_payload == payload


def test_validation_error_can_expose_diagnostic_artifacts(tmp_path):
    issue = ValidationIssue(
        rule_id="TEST.FAILURE",
        severity="error",
        message="invalid inventory",
    )
    result = ValidationRuleResult(
        rule_id=issue.rule_id,
        severity="error",
        applicability="applicable",
        checked_object_count=1,
        issues=(issue,),
    )
    phase = ValidationPhaseResult(
        phase_id="graph:full", kind="graph", rule_results=(result,)
    )
    report = ValidationReport(
        scenario_identity=("image", "path", 2050),
        store_generation=1,
        ruleset_version=1,
        certificate_key="certificate",
        rule_results=(result,),
        phase_results=(phase,),
    )
    error = PremiseValidationError(report)
    artifacts = premise.ChangeReportArtifacts(
        report_id="report",
        status="failed",
        workbook_path=tmp_path / "diagnostic.xlsx",
        details_path=None,
        scenario_identities=(("image", "path", 2050),),
        source_fingerprint="source",
        validation_certificate_keys=("certificate",),
    )

    error.attach_report_artifacts(artifacts)

    assert error.artifacts is artifacts
    assert str(artifacts.workbook_path) in str(error)


def test_validation_failure_automatically_builds_failed_workbook(tmp_path, monkeypatch):
    monkeypatch.chdir(tmp_path)
    source, final = _synthetic_stores()
    issue = ValidationIssue(
        rule_id="TEST.FAILURE", severity="error", message="invalid inventory"
    )
    result = ValidationRuleResult(
        rule_id=issue.rule_id,
        severity="error",
        applicability="applicable",
        checked_object_count=1,
        issues=(issue,),
    )
    phase = ValidationPhaseResult(
        phase_id="graph:full", kind="graph", rule_results=(result,)
    )
    report = ValidationReport(
        scenario_identity=("image", "path", 2050, ()),
        store_generation=final.generation,
        ruleset_version=1,
        certificate_key="failed-certificate",
        rule_results=(result,),
        phase_results=(phase,),
    )
    error = PremiseValidationError(report)
    scenario = {
        "model": "image",
        "pathway": "path",
        "year": 2050,
        "applied functions": ["electricity"],
        "_inventory_store": final,
    }
    database = object.__new__(NewDatabase)
    database.scenarios = [scenario]
    database._source_inventory_store = source
    database._compact_source_checkpoint = None
    database._validation_enabled = False
    database._validation_reports = {}
    database._validation_iam_fingerprints = {}
    database.generate_reports = True
    database._automatic_report_in_progress = False
    database._change_report_cache = None
    database.build_id = "1234567890abcdef"
    database.source = "source-db"
    database.source_type = "brightway"
    database.version = "3.12"
    database.system_model = "cutoff"

    database._generate_validation_diagnostic(error, scenario, final)

    assert error.artifacts.status == "failed"
    assert error.artifacts.workbook_path.is_file()
    assert str(error.artifacts.workbook_path) in str(error)
    workbook = openpyxl.load_workbook(error.artifacts.workbook_path)
    assert workbook["Overview"]["B3"].value == "failed"
    findings = workbook["Validation Findings"]
    assert any(
        findings.cell(row, 3).value == "TEST.FAILURE"
        for row in range(2, findings.max_row + 1)
    )


def test_export_failure_diagnostic_uses_invalid_runtime_inventory(monkeypatch):
    _, certified = _synthetic_stores()
    invalid_inventory = [
        _activity(
            "invalid",
            "invalid exporter activity",
            "GLO",
            [_production("invalid", "invalid exporter activity", "GLO")],
        )
    ]
    issue = ValidationIssue(
        rule_id="EXPORT.FAILURE", severity="error", message="invalid export graph"
    )
    result = ValidationRuleResult(
        rule_id=issue.rule_id,
        severity="error",
        applicability="applicable",
        checked_object_count=1,
        issues=(issue,),
    )
    phase = ValidationPhaseResult(
        phase_id="export:generic", kind="export", rule_results=(result,)
    )
    report = ValidationReport(
        scenario_identity=("image", "path", 2050, ()),
        store_generation=certified.generation,
        ruleset_version=1,
        certificate_key="certified-key",
        rule_results=(result,),
        phase_results=(phase,),
    )
    definition = {
        "model": "image",
        "pathway": "path",
        "year": 2050,
        "_inventory_store": certified,
        "_validation_report": report.to_dict(),
    }
    runtime = {
        "model": "image",
        "pathway": "path",
        "year": 2050,
        "database": invalid_inventory,
    }
    database = object.__new__(NewDatabase)
    database.inventory_backend = "compact"
    database._validation_reports = {}
    captured = {}

    def capture(error, scenario, store):
        captured.update(error=error, scenario=scenario, store=store)

    monkeypatch.setattr(database, "_generate_validation_diagnostic", capture)
    error = PremiseValidationError(report)

    database._handle_export_validation_error(definition, error, "brightway", runtime)

    assert captured["scenario"] is runtime
    assert len(captured["store"]) == 1
    assert captured["store"].activity(0)["name"] == "invalid exporter activity"
    assert error.report.phase_results[-1].phase_id == "export:brightway"
